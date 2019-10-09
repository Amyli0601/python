"""
============================
Author:柠檬班-木森
Time:2019/8/20
E-mail:3247119728@qq.com
Company:湖南零檬信息技术有限公司
============================
"""

import openpyxl


class CaseData(object):
    """测试用数据类"""

    def __init__(self, zip_obj):
        # 变量zip对象
        for i in list(zip_obj):
            # 通过反射机制设置属性
            # 将表头作为属性，值作为属性值
            setattr(self, i[0], i[1])


class CaseData2(object):
    pass


# [('表头'，值),(),()]
# 定义一个类专门用例读取excel中的数据

class ReadExcel(object):
    """读取excel中的用例数据"""

    def __init__(self, file_name, sheet_name):
        """
        :param file_name: excel文件名
        :param sheet_name: sheet表单名
        """
        self.file_name = file_name
        self.sheet_name = sheet_name

    def open(self):
        """打开工作薄和表单"""
        # # 打开文件，返回一个工作薄对象
        self.wb = openpyxl.load_workbook(self.file_name)
        # 通过工作薄，选择表单对象
        self.sh = self.wb[self.sheet_name]

    def read_data(self):
        """读取所有用例数据"""

        # 打开文件和表单
        self.open()
        # 按行获取所有的表格对象，每一行的内容放在一个元祖中，以列表的形式返回
        rows = list(self.sh.rows)
        # 创建一个列表cases,存放所有的用例数据
        cases = []
        # 获取表头
        titles1 = [r.value for r in rows[0]]
        # 遍历其他的数据行，和表头进行打包，转换为字典，放到cases这个列表中
        for row in rows[1:]:
            # 获取该行数据
            data = [r.value for r in row]
            # 和表头进行打包，转换为字典
            case = dict(zip(titles1, data))
            cases.append(case)
        # 将读取出来的数据进行返回
        return cases

    def read_data_obj(self):
        # 打开工作簿
        self.open()
        # 创建一个空的列表，用例存放所有的用例数据
        cases = []
        # 读取表单中的数据
        rows = list(self.sh.rows)
        # 读取表头
        print(rows[0])
        titles = []
        for r in rows[0]:
            titles.append(r.value)

        # 按行遍历
        for row in rows[1:]:
            # 读每一行数据
            case = []
            for r in row:
                case.append(r.value)
            zip_obj = zip(titles, case)
            # 将每一条用例的数据，存储为一个对象
            # 通过Case这个类来创建一个对象，传了一个参数，zip_obj
            case_data = CaseData(zip_obj)
            cases.append(case_data)
        # 将包含所有用例的列表cases进从返回
        return cases

    def read_data_dict2(self):
        self.open()
        cases = []
        max_row = self.sh.max_row
        for r in range(1, max_row + 1):
            case_dict = []
            case_dict['case_id'] = self.sh.cell(r, 1).value
            case_dict['excepted'] = self.sh.cell(r, 2).value
            case_dict['data'] = self.sh.cell(r, 3).value
            cases.append(cases)

    def read_data_obj2(self):
        self.open()
        cases = []
        # 获取最大行
        max_row = self.sh.max_row
        for r in range(1, max_row + 1):
            case_obj = CaseData2()
            case_obj.case_id = self.sh.cell(r, 1).value
            case_obj.excpted = self.sh.cell(r, 2).value
            case_obj.data = self.sh.cell(r, 3).value
            cases.append(cases)


if __name__ == '__main__':
    do_excel = ReadExcel('cases.xlsx', 'Sheet1')
    pass

    # do_excel.read_data_dict()
    # cases = do_excel.read_data_obj()
    # for i in cases:
    #     print(i.case_id)
    #     print(i.data)
    # cases2 = do_excel.read_data()
    # print(cases2)

# a = {"data":(11,22,33),"excepted":{"nn":"预期结果"}}
# c = CaseData(data=(11,22,33),excepted={"nn":"预期结果"})
#
# print(a['data'])
# print(c.data)
