"""
===============
author:Administrator
time:15:56
E-mail:1223607348@qq.com
===============
"""
import openpyxl
# 打开文件，返回一个工作簿对象
wb = openpyxl.load_workbook("cases.xlsx")
# 通过工作簿，选择表单对象
sh01 = wb['test01']
# # 读取表格数据:单个
# cell01=eval(sh01.cell(row=3,column=2).value)
# 读取表格数据:多个
# 读取多个数据方式一：
# 获取总行数
max_row =sh01.max_row
# print(max_row)
# 获取总列数
max_column = sh01.max_column
# print(max_column)
cases =[]
for i in range(1,max_row+1):
    case_dict={}
    # 判断是否为表头，创建空列表存储表头数据
    if i==1:
        titles = []
        for j in range(1,max_column+1):
            res =sh01.cell(row=i,column=j).value
            titles.append(res)
    else:
        # 如果不是表头，创建空列表存储该行数据
        datas = []
        for j in range(1,max_column+1):
            res =sh01.cell(row=i,column=j).value
            datas.append(res)
        # 将行和列的数据打包
        case_zip = zip(titles,datas)
        # 将打包后的数据转化为一个字典
        case = dict(case_zip)
        # 将每条数据存入cases列表中
        cases.append(case)
# print(cases)
# 读取多个数据方式二：
# 按行获取所有的表格对象，每行内容放在一个
rows = list(sh01.rows)
# 获取表头方法一:
titles =[]
for row in rows[0]:
    titles.append(row.value)
# 获取表头方法二:
# 创建空列表存储读取数据
cases = []
# 列表推导式
titles1 = [row.value for row in rows[0] ]
# 获取最大值
# 遍历非表头的行，和表头一起打包为一个字典
for row in rows[1:]:
    data = [r.value for r in row]
    case = dict(zip(titles1,data))
    cases.append(case)
print(cases)




# # 往表格中写数据
# cell02=sh01.cell(row=8,column=3)
# cell02.value = '测试'
# # 保存数据
# wb.save('cases.xlsx')