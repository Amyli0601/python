"""
===============
author:Administrator
time:14:55
E-mail:1223607348@qq.com
===============
"""
'''
openpyxl:仅支持xlxs格式文件的读取和编辑
先导入模块，安装命令:pip install openpyxl

Excel中的三大对象：
1、WorkBook:工作簿对象
2、Sheet:表单对象
3、Cell:表格对象(单元格)

Excel读取出来的数据均为字符串或int
'''
import openpyxl
# 打开文件，返回一个工作簿对象
wb = openpyxl.load_workbook('cases.xlsx')
# print(wb)
# 通过工作簿选中表单对象
sh = wb['cases']
# 通过表单对象读取表格中的数据:
# # 1、获取单个数据
# cell01 = sh.cell(row=1,column=1).value
# cell02 = eval(sh.cell(row=2,column=2).value)
# cell03 = eval(sh.cell(row=2,column=3).value)
# print(cell01,type(cell01))
# print(cell02,type(cell02))
# print(cell03,type(cell03))
# 2、读取多个数据：
# （1）已知行和列
# cases = []
# for i in range(1,6):
#         # 判断是否为表头
#         if i==1:
#             titles = []
#             for j in range(1, 4):
#                 data = sh.cell(row=i,column=j).value
#                 titles.append(data)
#         else:
#             # 非表头，创建空列表存储数据
#             datas = []
#             for j in range(1, 4):
#                 data = sh.cell(row=i, column=j).value
#                 datas.append(data)
#             # 将表头和数据打包为字典，每个字典为一个用例数据
#             cases_zip = zip(titles,datas)
#             case = dict(cases_zip)
#             # 将用例数据添加到cases列表中
#             cases.append(case)
# print(cases)
# # （2）未知行和列:
# # 获取总行数
# max_row = sh.max_row
# # print(max_row)
# # 获取总列数
# max_column = sh.max_column
# # print(max_column)
# cases = []
# for i in range(1,max_row+1):
#         # 判断是否为表头
#         if i==1:
#             titles = []
#             for j in range(1, max_column+1):
#                 data = sh.cell(row=i,column=j).value
#                 titles.append(data)
#         else:
#             # 非表头，创建空列表存储数据
#             datas = []
#             for j in range(1, max_column+1):
#                 data = sh.cell(row=i, column=j).value
#                 datas.append(data)
#             # 将表头和数据打包为字典，每个字典为一个用例数据
#             cases_zip = zip(titles,datas)
#             case = dict(cases_zip)
#             # 将用例数据添加到cases列表中
#             cases.append(case)
# print(cases)
# (3)
# 按行获取所有的表格对象，每一行放在一个列表中
rows = list(sh.rows)
# # 获取表头：
# # 方法一：常规方法
# titles = []
# for row in rows[0]:
#     titles.append(row.value)
# 方法二：列表推导式
# 创建一个空列表cases，存放所有的用例数据
cases = []
# 获取表头
titles1 = [row.value for row in rows[0]]
# 遍历非表头的行,和表头打包一起打包成字典，添加到cases列表中
for row in rows[1:]:
    data = [r.value for r in row]
    case = dict(zip(titles1,data))
    cases.append(case)
print(cases)

# # 往表格中写数据
# cell00 = sh.cell(row=1,column=4)
# cell00.value = "result"
#
# #  保存数据（写完数据保存才会生效）
# wb.save('cases.xlsx')

