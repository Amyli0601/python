"""
============================
author:LiuYu
time:2019/8/23
E-mail:253342381@qq.com
============================
"""

import openpyxl
from openpyxl import load_workbook

class DoExcsl:

    def __init__(self,file_name,Sheet):
        self.file_name = file_name
        self.Sheet = Sheet
    def login_excel(self):  #读取充值模块的用例
        # case_id = ReadConfig(project_path11.conf_path).get_data("My_log", "case_id")
        # print(case_id)
        wb = load_workbook(self.file_name)  # 打开表格
        sheet = wb[self.Sheet]
        # 储存所有的数据 # 注意这里获取的是整数
        text_data = []
        for i in range(2, sheet.max_row + 1):
            sub_data = {}
            sub_data["Case_id"] = sheet.cell(i, 1).value
            sub_data["Module"] = sheet.cell(i,2).value
            sub_data["Params"] = sheet.cell(i, 3).value
            sub_data["cell"] = sheet.cell(i, 4).value
            sub_data["Method"] = sheet.cell(i, 5).value
            sub_data["Params"] = sheet.cell(i,6).value
            sub_data["ExpectedResult"] = sheet.cell(i, 7).value
            text_data.append(sub_data)
        # final_data = []
        # if self.case_conf == "all":  # 配置文件控制用例如果是列表，那就获取列表里面指定id对应的数据
        #     final_data = text_data  # 把测试用例赋值给final_data这个变量
        # else:
        #     for i in self.case_conf:  # 遍历case_id里面的值
        #         final_data.append(text_data[i-1])
        return text_data
    def write_excel(self, row, column, values):
        #row代表写入的行
        #column代表写入的列
        #values代表写入值
            '''在指定的单元格写入指定的数据，并保存到当前Excel'''
            try:  # 输入的变量可能导致异常， 例如文件名、表单名不存在，因此加入异常处理
                wb = load_workbook('case.xlsx')
                sheet = wb["login"]
                sheet.cell(row, column).value = values
                wb.save('case.xlsx')
            except KeyError as e:
                print("您的输入有误，报错为：{}".format(e))  # 打印报错信息
                exit()  # 退出程序，注意这里不能用break，break用于for和while循环中

if __name__ == '__main__':
    a = DoExcsl('case.xlsx','liuyu').login_excel()
    print(a)
    # cad = open_method('case.xlsx','liuyu')
    # cases = cad.test_case()
    # print(cases)














# work = openpyxl.load_workbook('case.xlsx')
# sh = work['liuyu']
#
# rows = list(sh.rows)
#
# cases =[]
# title1 = [row.value for row in rows[0]]
# print(title1)
#
# for row in rows[1:]:
#     case = [r.value for r in row]
#     data = dict(zip(title1,case))
#     cases.append(data)
# print(cases)



# title1 = [row.value for row in rows[0]]
# print(title1)
# # for row in rows[0]:
# #     print(list(row.value))
# for row in rows[1:]:
#     data = [r.value for r in row]
#     case =dict(zip(title1,data))
#     cases.append(case)
# print(cases)
