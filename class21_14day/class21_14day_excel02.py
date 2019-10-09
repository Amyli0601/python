"""
============================
Author:柠檬班-木森
Time:2019/8/20
E-mail:3247119728@qq.com
Company:湖南零檬信息技术有限公司
============================
"""

"""
openpyxl：
工作簿workbook:

表单sheet:


表格cell:
"""
import openpyxl

# # 打开文件，返回一个工作薄对象
wb = openpyxl.load_workbook("cases1.xlsx")

# 通过工作薄，选择表单对象
sh = wb['register']


# 获取行总数
max_row = sh.max_row
print(max_row)

# 获取列总数
max_column = sh.max_column
print(max_column)


# 读取文件中的数据
cases = []
for r in range(1, max_row+1):
    case_dict = {}
    # 判断是否是表头（是否是第一行）
    if r == 1:
        titles = []
        for c in range(1, max_column+1):
            data = sh.cell(row=r, column=c).value
            titles.append(data)
    else:
        # 如果不是表头，创建一个空列表来存储改行数据
        datas = []
        for c in range(1, max_column+1):
            data = sh.cell(row=r, column=c).value
            datas.append(data)
        # 将表头和改行数据进行打包
        case_zip = zip(titles, datas)
        # 将打包后的zip对象转换为字典，每个字典就是一条用例数据
        case = dict(case_zip)
        # 将该条用例数据，加入cases这个列表中
        cases.append(case)

print(cases)
