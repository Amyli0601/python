"""
============================
Author:柠檬班-木森
Time:2019/8/20
E-mail:3247119728@qq.com
Company:湖南零檬信息技术有限公司
============================
"""

"""
openpyxl：
工作簿workbook:

表单sheet:


表格cell:
"""
import openpyxl

# 打开文件，返回一个工作薄对象
wb = openpyxl.load_workbook("cases.xlsx")

# 通过工作薄，选择表单对象
sh = wb['musen']



c11 = sh.cell(row=1, column=1).value
c12 = sh.cell(row=1, column=2).value
c13 = sh.cell(row=1, column=3).value
c14 = sh.cell(row=1, column=4).value
c21 = sh.cell(row=2, column=1).value
c22 = sh.cell(row=2, column=2).value
c23 = sh.cell(row=2, column=3).value
c24 = sh.cell(row=2, column=4).value



# 往表格中写数据
# c12 = sh.cell(row=1, column=2)
# c12.value = "python第一"

# 保存数据（写完之后一定保存才会生效）
# wb.save('cases.xlsx')